# -*- coding: utf-8 -*-
"""
Xuat toan bo bao gia da thu thap tu website ra file Excel, dung DUNG cau truc
sheet Tong_hop_bao_gia da duoc xay dung thu cong truoc do (bang chi tiet A-O,
vung tham chieu Q-T, bang tong hop/chot gia du toan ben duoi) - de NCC noi tiep
dung 1 quy trinh voi cac dot truoc.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.worksheet.datavalidation import DataValidation

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF")
THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

DETAIL_START_ROW = 5
DETAIL_END_ROW = 5 + 300  # du cho nhieu bao gia hon cac dot truoc


def build_workbook(vendor_by_id, all_items, all_maint, categories):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tong_hop_bao_gia"

    ws["A1"] = "TỔNG HỢP BÁO GIÁ - GÓI THẦU TBYT (xuất tự động từ Cổng thu thập báo giá trực tuyến)"
    ws["A2"] = (
        "Nền XANH = NCC chào giá THẤP NHẤT | Nền ĐỎ = NCC chào giá CAO NHẤT | "
        "Cột N,O tham khảo (bảo trì mở rộng) | Cột M/O bảng tổng hợp bên dưới do Bệnh viện tự quyết định"
    )

    # 1. TIÊU ĐỀ BẢNG CHI TIẾT
    headers = [
        "Mã YCBG", "Danh mục thiết bị y tế", "Tên nhà cung cấp", "Mã số thuế", "Địa chỉ công ty",
        "Ký/mã/nhãn hiệu/Model", "Hãng sản xuất", "Xuất xứ", "Số lượng", "ĐVT",
        "Đơn giá\n(đã bao gồm thuế, phí, lệ phí)", "Thành tiền", "Ngày báo giá",
        "Giá bảo trì/bảo hành\nmở rộng - gói 12 tháng\n(1 thiết bị, đã VAT)",
        "Thành tiền bảo trì\nmở rộng (toàn bộ SL)"
    ]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        c.border = BORDER

    # 2. VÙNG THAM CHIẾU DANH MỤC (Cột Q-T)
    ws["Q4"] = "Mã YCBG"
    ws["R4"] = "Danh mục thiết bị y tế (Vùng tham chiếu - KHÔNG XOÁ)"
    ws["S4"] = "ĐVT"
    ws["T4"] = "SL yêu cầu"
    for c in ["Q4", "R4", "S4", "T4"]:
        ws[c].font = HEADER_FONT
        ws[c].fill = HEADER_FILL
        ws[c].alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    for i, cat in enumerate(categories):
        r = 5 + i
        ma_code = cat.get("ma", f"TB {i+1:02d}")
        if len(ma_code) == 4 and ma_code.upper().startswith("TB") and ma_code[2:].isdigit():
            formatted_cat_ma = f"TB {ma_code[2:]}"
        else:
            formatted_cat_ma = ma_code
        ws.cell(row=r, column=17, value=formatted_cat_ma)  # Cột Q
        ws.cell(row=r, column=18, value=cat["ten"])         # Cột R
        ws.cell(row=r, column=19, value=cat["dvt"])         # Cột S
        ws.cell(row=r, column=20, value=cat["sl"])          # Cột T
    ref_last_row = 4 + len(categories)

    # 3. ĐIỀN DỮ LIỆU TỪ CÁC NHÀ THẦU ĐÃ NỘP (Truy xuất dữ liệu an toàn)
    r = DETAIL_START_ROW
    for item in all_items:
        v = vendor_by_id.get(item["vendor_id"])
        if not v:
            continue

        ma_raw = item["ma_danh_muc"] or ""
        if len(ma_raw) == 4 and ma_raw.upper().startswith("TB") and ma_raw[2:].isdigit():
            ma_display = f"TB {ma_raw[2:]}"
        else:
            ma_display = ma_raw

        ten_ncc = v["ten_ncc"] or ""
        mst = v["mst"] if ("mst" in v.keys() and v["mst"]) else ""
        dia_chi = v["dia_chi"] if ("dia_chi" in v.keys() and v["dia_chi"]) else ""
        model = item["model"] or ""
        hang_sx = item["hang_sx"] or ""
        xuat_xu = item["xuat_xu"] or ""
        don_gia = item["don_gia"]
        ngay_bao_gia = item["ngay_bao_gia"] if ("ngay_bao_gia" in item.keys() and item["ngay_bao_gia"]) else ""

        # Cột A: Hiển thị Mã thiết bị TB 01, TB 02...
        ws.cell(row=r, column=1, value=ma_display)
        ws.cell(row=r, column=2, value=next((c["ten"] for c in categories if c["ma"] == item["ma_danh_muc"]), item["ma_danh_muc"]))
        ws.cell(row=r, column=3, value=ten_ncc)
        ws.cell(row=r, column=4, value=mst)
        ws.cell(row=r, column=5, value=dia_chi)
        ws.cell(row=r, column=6, value=model)
        ws.cell(row=r, column=7, value=hang_sx)
        ws.cell(row=r, column=8, value=xuat_xu)
        ws.cell(row=r, column=9, value=f'=IF($B{r}="","",VLOOKUP($B{r},$R$5:$T${ref_last_row},3,0))')
        ws.cell(row=r, column=10, value=f'=IF($B{r}="","",VLOOKUP($B{r},$R$5:$T${ref_last_row},2,0))')
        ws.cell(row=r, column=11, value=don_gia)
        ws.cell(row=r, column=12, value=f'=IF(OR($I{r}="",$K{r}=""),"",$I{r}*$K{r})')
        ws.cell(row=r, column=13, value=ngay_bao_gia)

        maint_match = next((m for m in all_maint if m["vendor_id"] == item["vendor_id"] and m["ma_danh_muc"] == item["ma_danh_muc"]), None)
        if maint_match:
            ws.cell(row=r, column=14, value=maint_match["don_gia_baotri"])
        ws.cell(row=r, column=15, value=f'=IF(OR($I{r}="",$N{r}=""),"",$I{r}*$N{r})')
        r += 1

    # Điền công thức rỗng cho các dòng còn lại đến DETAIL_END_ROW
    for rr in range(r, DETAIL_END_ROW + 1):
        ws.cell(row=rr, column=9, value=f'=IF($B{rr}="","",VLOOKUP($B{rr},$R$5:$T${ref_last_row},3,0))')
        ws.cell(row=rr, column=10, value=f'=IF($B{rr}="","",VLOOKUP($B{rr},$R$5:$T${ref_last_row},2,0))')
        ws.cell(row=rr, column=12, value=f'=IF(OR($I{rr}="",$K{rr}=""),"",$I{rr}*$K{rr})')
        ws.cell(row=rr, column=15, value=f'=IF(OR($I{rr}="",$N{rr}=""),"",$I{rr}*$N{rr})')

    last_detail_row = DETAIL_END_ROW

    dv = DataValidation(type="list", formula1=f"$R$5:$R${ref_last_row}", allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"B{DETAIL_START_ROW}:B{last_detail_row}")

    # 4. TÔ MÀU MIN / MAX (Nền XANH = Giá thấp nhất, Nền ĐỎ = Giá cao nhất)
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    b1 = DETAIL_START_ROW
    e1 = last_detail_row

    rule_min = Rule(
        type="expression",
        formula=[f'AND($B{b1}<>"",COUNTIFS($B${b1}:$B${e1},$B{b1})>1,$K{b1}=MINIFS($K${b1}:$K${e1},$B${b1}:$B${e1},$B{b1}))'],
        dxf=DifferentialStyle(fill=green_fill),
        stopIfTrue=False,
    )
    rule_max = Rule(
        type="expression",
        formula=[
            f'AND($B{b1}<>"",COUNTIFS($B${b1}:$B${e1},$B{b1})>1,$K{b1}=MAXIFS($K${b1}:$K${e1},$B${b1}:$B${e1},$B{b1}),'
            f'MINIFS($K${b1}:$K${e1},$B${b1}:$B${e1},$B{b1})<>MAXIFS($K${b1}:$K${e1},$B${b1}:$B${e1},$B{b1}))'
        ],
        dxf=DifferentialStyle(fill=red_fill),
        stopIfTrue=False,
    )
    # Áp dụng tô màu toàn bộ dòng cho dải ô A đến O
    ws.conditional_formatting.add(f"A{DETAIL_START_ROW}:O{last_detail_row}", rule_min)
    ws.conditional_formatting.add(f"A{DETAIL_START_ROW}:O{last_detail_row}", rule_max)

    # 5. BẢNG TỔNG HỢP GIÁ & CHỐT GIÁ DỰ TOÁN
    sum_start = last_detail_row + 3
    ws.cell(row=sum_start - 2, column=1, value="BẢNG TỔNG HỢP GIÁ & CHỐT GIÁ DỰ TOÁN THEO TỪNG DANH MỤC").font = Font(bold=True, size=12)
    hdr2 = [
        "STT", "Danh mục thiết bị y tế", "ĐVT", "SL\nyêu cầu", "Số báo\ngiá nhận\nđược",
        "Giá thấp nhất\n(đã VAT)", "NCC chào\ngiá thấp nhất", "Giá cao nhất\n(đã VAT)",
        "NCC chào\ngiá cao nhất", "Giá trung bình\n(đã VAT)", "% Chênh\nlệch", "Cảnh báo",
        "Giá dự toán\nCHỐT (đã VAT)", "Thành tiền\ndự toán", "Căn cứ chọn giá"
    ]
    for i, h in enumerate(hdr2, start=1):
        c = ws.cell(row=sum_start, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    for i, cat in enumerate(categories):
        r = sum_start + 1 + i
        refrow = 5 + i
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=f"=R{refrow}")
        ws.cell(row=r, column=3, value=f"=S{refrow}")
        ws.cell(row=r, column=4, value=f"=T{refrow}")
        ws.cell(row=r, column=5, value=f'=COUNTIF($B${DETAIL_START_ROW}:$B${last_detail_row},$B{r})')
        ws.cell(row=r, column=6, value=f'=IF($E{r}=0,"",MINIFS($K${DETAIL_START_ROW}:$K${last_detail_row},$B${DETAIL_START_ROW}:$B${last_detail_row},$B{r}))')
        ws.cell(row=r, column=7, value=(
            f'=IF($E{r}=0,"",INDEX($C${DETAIL_START_ROW}:$C${last_detail_row},'
            f'SUMPRODUCT(($B${DETAIL_START_ROW}:$B${last_detail_row}=$B{r})*($K${DETAIL_START_ROW}:$K${last_detail_row}=$F{r})*'
            f'ROW($B${DETAIL_START_ROW}:$B${last_detail_row}))-{DETAIL_START_ROW - 1}))'
        ))
        ws.cell(row=r, column=8, value=f'=IF($E{r}=0,"",MAXIFS($K${DETAIL_START_ROW}:$K${last_detail_row},$B${DETAIL_START_ROW}:$B${last_detail_row},$B{r}))')
        ws.cell(row=r, column=9, value=(
            f'=IF($E{r}=0,"",INDEX($C${DETAIL_START_ROW}:$C${last_detail_row},'
            f'SUMPRODUCT(($B${DETAIL_START_ROW}:$B${last_detail_row}=$B{r})*($K${DETAIL_START_ROW}:$K${last_detail_row}=$H{r})*'
            f'ROW($B${DETAIL_START_ROW}:$B${last_detail_row}))-{DETAIL_START_ROW - 1}))'
        ))
        ws.cell(row=r, column=10, value=f'=IF($E{r}=0,"",AVERAGEIF($B${DETAIL_START_ROW}:$B${last_detail_row},$B{r},$K${DETAIL_START_ROW}:$K${last_detail_row}))')
        ws.cell(row=r, column=11, value=f'=IF(OR($E{r}<2,$J{r}=""),"",($H{r}-$F{r})/$J{r})')
        ws.cell(row=r, column=12, value=f'=IF($K{r}="","",IF($K{r}>0.2,"⚠ Chênh lệch cao","OK"))')
        ws.cell(row=r, column=14, value=f'=IF(OR($D{r}="",$M{r}=""),"",$D{r}*$M{r})')

    warn_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    l_first = sum_start + 1
    l_last = sum_start + len(categories)
    ws.conditional_formatting.add(
        f"L{l_first}:L{l_last}",
        Rule(type="expression", formula=[f'$L{l_first}="⚠ Chênh lệch cao"'], dxf=DifferentialStyle(fill=warn_fill)),
    )

    total_row = sum_start + 1 + len(categories)
    ws.cell(row=total_row, column=2, value="TỔNG DỰ TOÁN GÓI THẦU").font = Font(bold=True)
    ws.cell(row=total_row, column=14, value=f"=SUM(N{sum_start+1}:N{total_row-1})").font = Font(bold=True)

    # 6. ĐỘ RỘNG CÁC CỘT
    widths = {
        1: 10, 2: 30, 3: 26, 4: 15, 5: 32, 6: 16, 7: 18, 8: 12, 9: 10, 10: 8,
        11: 16, 12: 16, 13: 12, 14: 16, 15: 16,
        17: 10, 18: 34, 19: 10, 20: 10
    }
    for col, w in widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    return wb
