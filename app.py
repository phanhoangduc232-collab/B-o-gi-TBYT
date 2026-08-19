# -*- coding: utf-8 -*-
"""
Cong thu thap bao gia truc tuyen - Benh vien E
Goi thau: Mua sam TBYT 2026 (Quy PTHDSN, Dot III) - YCBG so 2717/YCBG-BVE

Nha cung cap tu dang ky bang Ten cong ty + Ma so thue + mat khau tu dat,
chi xem/sua duoc du lieu cua chinh minh trong thoi gian mo (SUBMISSION_OPEN_AT -> SUBMISSION_CLOSE_AT).
"""
import os
import io
import uuid
from datetime import datetime

from flask import (
    Flask, render_template, request, redirect, url_for, session, flash,
    send_from_directory, abort, g
)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

try:
    from zoneinfo import ZoneInfo
    TZ = ZoneInfo("Asia/Ho_Chi_Minh")
except Exception:  # pragma: no cover
    from datetime import timezone, timedelta
    TZ = timezone(timedelta(hours=7))

from db import get_db, init_db, DB_PATH
from categories import CATEGORIES, CATEGORY_BY_MA, FILE_TAGS

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.environ.get("DATA_DIR", os.path.join(BASE_DIR, "instance"))
UPLOAD_DIR = os.path.join(DATA_DIR, "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

ALLOWED_EXT = {"pdf", "doc", "docx", "xls", "xlsx", "jpg", "jpeg", "png", "rar", "zip"}
MAX_CONTENT_LENGTH_MB = int(os.environ.get("MAX_CONTENT_LENGTH_MB", "25"))

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "doi-chuoi-nay-truoc-khi-deploy-that")
app.config["MAX_CONTENT_LENGTH"] = MAX_CONTENT_LENGTH_MB * 1024 * 1024

# Khoi tao DB ngay khi module duoc nap (ke ca khi chay qua gunicorn, khong chi qua __main__)
init_db()

ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "doi-mat-khau-admin")
GOI_THAU_TEN = os.environ.get(
    "GOI_THAU_TEN",
    "Mua sắm Thiết bị y tế bằng Nguồn Quỹ Phát triển Hoạt động sự nghiệp và các nguồn vốn hợp pháp khác của Bệnh viện E năm 2026 (Đợt III)",
)
YCBG_SO = os.environ.get("YCBG_SO", "2717/YCBG-BVE ngày 14/7/2026")


def _parse_dt(env_name, default_iso):
    raw = os.environ.get(env_name, default_iso)
    dt = datetime.fromisoformat(raw)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=TZ)
    return dt


SUBMISSION_OPEN_AT = _parse_dt("SUBMISSION_OPEN_AT", "2026-08-01T10:00:00")
SUBMISSION_CLOSE_AT = _parse_dt("SUBMISSION_CLOSE_AT", "2026-08-15T16:00:00")


def now_vn():
    return datetime.now(TZ)


def window_status():
    n = now_vn()
    if n < SUBMISSION_OPEN_AT:
        return "before"
    if n > SUBMISSION_CLOSE_AT:
        return "after"
    return "open"


@app.template_filter("vn_dt")
def vn_dt(value):
    """Dinh dang chuoi ISO datetime luu trong DB thanh dd/mm/yyyy HH:MM de hien thi."""
    if not value:
        return ""
    try:
        dt = datetime.fromisoformat(value)
        return dt.strftime("%d/%m/%Y %H:%M")
    except (ValueError, TypeError):
        return value


@app.context_processor
def inject_globals():
    return dict(
        goi_thau_ten=GOI_THAU_TEN,
        ycbg_so=YCBG_SO,
        open_at=SUBMISSION_OPEN_AT,
        close_at=SUBMISSION_CLOSE_AT,
        window_status=window_status(),
        is_admin=session.get("is_admin", False),
    )


# ---------------------------------------------------------------- helpers --

def current_vendor():
    vid = session.get("vendor_id")
    if not vid:
        return None
    db = get_db()
    row = db.execute("SELECT * FROM vendors WHERE id=?", (vid,)).fetchone()
    db.close()
    return row


def require_vendor():
    v = current_vendor()
    if not v:
        return None
    return v


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXT


# ------------------------------------------------------------------ routes --

@app.route("/")
def index():
    if current_vendor():
        return redirect(url_for("form"))
    return redirect(url_for("dangky"))


@app.route("/dangky", methods=["GET", "POST"])
def dangky():
    status = window_status()
    if request.method == "POST":
        if status == "before":
            flash("Chưa đến thời gian mở nhận báo giá.", "danger")
            return redirect(url_for("dangky"))

        ten_ncc = request.form.get("ten_ncc", "").strip()
        mst = request.form.get("mst", "").strip()
        mat_khau = request.form.get("mat_khau", "")

        if not ten_ncc or not mst or not mat_khau:
            flash("Vui lòng nhập đầy đủ Tên nhà cung cấp, Mã số thuế và Mật khẩu.", "danger")
            return redirect(url_for("dangky"))

        db = get_db()
        row = db.execute("SELECT * FROM vendors WHERE mst=?", (mst,)).fetchone()

        if row is None:
            # dang ky moi - chi cho phep trong thoi gian mo
            if status == "after":
                flash("Đã hết thời hạn nhận báo giá, không thể đăng ký mới.", "danger")
                db.close()
                return redirect(url_for("dangky"))
            pw_hash = generate_password_hash(mat_khau)
            cur = db.execute(
                "INSERT INTO vendors (ten_ncc, mst, password_hash, created_at) VALUES (?,?,?,?)",
                (ten_ncc, mst, pw_hash, now_vn().isoformat()),
            )
            db.commit()
            vendor_id = cur.lastrowid
            db.close()
            session.clear()
            session["vendor_id"] = vendor_id
            flash("Đăng ký thành công. Vui lòng nhập báo giá của quý công ty.", "success")
            return redirect(url_for("form"))
        else:
            # da ton tai MST nay -> yeu cau dung mat khau da dat truoc do de dang nhap lai
            if check_password_hash(row["password_hash"], mat_khau):
                db.close()
                session.clear()
                session["vendor_id"] = row["id"]
                return redirect(url_for("form"))
            else:
                db.close()
                flash(
                    "Mã số thuế này đã được đăng ký trước đó. Mật khẩu không đúng - "
                    "vui lòng dùng đúng mật khẩu quý công ty đã tự đặt khi đăng ký lần đầu.",
                    "danger",
                )
                return redirect(url_for("dangky"))

    return render_template("dangky.html", status=status)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("dangky"))


@app.route("/form")
def form():
    v = require_vendor()
    if not v:
        return redirect(url_for("dangky"))
    status = window_status()
    db = get_db()
    items = db.execute(
        "SELECT * FROM quote_items WHERE vendor_id=? ORDER BY id", (v["id"],)
    ).fetchall()
    maint = db.execute(
        "SELECT * FROM maintenance_items WHERE vendor_id=? ORDER BY id", (v["id"],)
    ).fetchall()
    files = db.execute(
        "SELECT * FROM files WHERE vendor_id=? ORDER BY id", (v["id"],)
    ).fetchall()
    db.close()
    readonly = status == "after"
    return render_template(
        "form.html",
        vendor=v,
        items=items,
        maint=maint,
        files=files,
        categories=CATEGORIES,
        category_by_ma=CATEGORY_BY_MA,
        file_tags=FILE_TAGS,
        readonly=readonly,
        status=status,
    )

@app.route("/form/company", methods=["POST"])
def form_company():
    v = require_vendor()
    if not v:
        return redirect(url_for("dangky"))
    if window_status() == "after":
        flash("Đã hết hạn, không thể chỉnh sửa.", "danger")
        return redirect(url_for("form"))
    db = get_db()
    db.execute(
        "UPDATE vendors SET dia_chi=?, nguoi_lien_he=?, sdt=?, email=?, submitted_at=? WHERE id=?",
        (
            request.form.get("dia_chi", "").strip(),
            request.form.get("nguoi_lien_he", "").strip(),
            request.form.get("sdt", "").strip(),
            request.form.get("email", "").strip(),
            now_vn().isoformat(),
            v["id"],
        ),
    )
    db.commit()
    db.close()
    flash("Đã lưu thông tin liên hệ.", "success")
    return redirect(url_for("form"))

import openpyxl

@app.route("/form/import-excel", methods=["POST"])
def import_excel():
    v = require_vendor()
    if not v:
        return redirect(url_for("dangky"))
    if window_status() == "after":
        flash("Đã hết thời hạn nhận báo giá, không thể nhập thêm.", "danger")
        return redirect(url_for("form"))

    file = request.files.get("excel_file")
    if not file or file.filename == "":
        flash("Vui lòng chọn file Excel để tải lên.", "danger")
        return redirect(url_for("form"))

    if not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        flash("Định dạng file không hợp lệ! Vui lòng chỉ tải lên file Excel (.xlsx, .xls).", "danger")
        return redirect(url_for("form"))

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        
        # 1. Tìm đúng sheet Phụ lục I Báo giá
        target_sheet_name = None
        for name in wb.sheetnames:
            if "Phụ lục I" in name or "Báo giá" in name:
                target_sheet_name = name
                break
        if not target_sheet_name:
            target_sheet_name = wb.sheetnames[0]

        ws = wb[target_sheet_name]
        db = get_db()
        imported_count = 0

        # 2. Quét thông minh từ dòng 4 đến hết bảng dữ liệu (tối đa dòng 100)
        for row in range(4, min(ws.max_row + 1, 100)):
            ma_raw = ws.cell(row=row, column=1).value  # Cột A: STT / Mã trong YCBG
            if not ma_raw:
                continue

            ma_str = str(ma_raw).strip()
            # Bỏ qua các dòng tiêu đề hoặc ghi chú
            if "STT" in ma_str or "Báo giá" in ma_str or "cam kết" in ma_str.lower():
                continue

            # Xóa dấu cách để khớp chuẩn (Ví dụ: "TB 01" -> "TB01")
            ma_clean = ma_str.replace(" ", "").upper()

            # Khớp với mã danh mục của gói thầu
            matched_code = None
            for c_code in CATEGORY_BY_MA.keys():
                if c_code.replace(" ", "").upper() == ma_clean:
                    matched_code = c_code
                    break

            if not matched_code:
                continue

            # Lấy các thông tin thiết bị
            ten_tm = str(ws.cell(row=row, column=2).value or "").strip()     # Cột B: Danh mục TBYT
            model = str(ws.cell(row=row, column=3).value or "").strip()      # Cột C: Model
            hang_sx = str(ws.cell(row=row, column=4).value or "").strip()    # Cột D: Hãng SX
            xuat_xu = str(ws.cell(row=row, column=6).value or "").strip()    # Cột F: Xuất xứ

            # Cột I: Đơn giá đã bao gồm thuế phí (Cột 9)
            don_gia_raw = ws.cell(row=row, column=9).value
            don_gia = None
            if don_gia_raw is not None:
                try:
                    # Làm sạch dấu phẩy/chấm nếu lưu dạng chuỗi
                    val_clean = str(don_gia_raw).replace(",", "").replace(".", "").strip()
                    don_gia = int(float(val_clean)) if val_clean else None
                except ValueError:
                    don_gia = None

            # Xóa dòng cũ nếu nhà thầu import lại cùng một mã thiết bị
            db.execute(
                "DELETE FROM quote_items WHERE vendor_id=? AND ma_danh_muc=?",
                (v["id"], matched_code),
            )

            # Thêm dòng báo giá vào database
            db.execute(
                """INSERT INTO quote_items
                   (vendor_id, ma_danh_muc, ten_thuong_mai, model, hang_sx, xuat_xu, don_gia, ngay_bao_gia, created_at)
                   VALUES (?,?,?,?,?,?,?,?,?)""",
                (
                    v["id"],
                    matched_code,
                    ten_tm,
                    model,
                    hang_sx,
                    xuat_xu,
                    don_gia,
                    now_vn().strftime("%d/%m/%Y"),
                    now_vn().isoformat(),
                ),
            )
            imported_count += 1

        db.commit()
        db.close()

        if imported_count > 0:
            flash(f"✅ Đã tự động đọc và nhập thành công {imported_count} danh mục thiết bị từ file Excel!", "success")
        else:
            flash("Không tìm thấy dòng thiết bị nào hợp lệ trong file Excel. Vui lòng kiểm tra lại đúng mẫu!", "warning")

    except Exception as e:
        flash(f"Lỗi khi đọc file Excel: {str(e)}", "danger")

    return redirect(url_for("form"))
@app.route("/form/item/add", methods=["POST"])
def item_add():
    v = require_vendor()
    if not v:
        return redirect(url_for("dangky"))
    if window_status() == "after":
        flash("Đã hết hạn, không thể chỉnh sửa.", "danger")
        return redirect(url_for("form"))

    ma = request.form.get("ma_danh_muc")
    if ma not in CATEGORY_BY_MA:
        flash("Danh mục không hợp lệ.", "danger")
        return redirect(url_for("form"))

    don_gia_raw = request.form.get("don_gia", "").replace(",", "").replace(".", "").strip()
    try:
        don_gia = int(don_gia_raw) if don_gia_raw else None
    except ValueError:
        don_gia = None

    db = get_db()
    db.execute(
        """INSERT INTO quote_items
           (vendor_id, ma_danh_muc, ten_thuong_mai, model, hang_sx, xuat_xu, don_gia, ngay_bao_gia, created_at)
           VALUES (?,?,?,?,?,?,?,?,?)""",
        (
            v["id"],
            ma,
            request.form.get("ten_thuong_mai", "").strip(),
            request.form.get("model", "").strip(),
            request.form.get("hang_sx", "").strip(),
            request.form.get("xuat_xu", "").strip(),
            don_gia,
            request.form.get("ngay_bao_gia", "").strip(),
            now_vn().isoformat(),
        ),
    )
    db.execute("UPDATE vendors SET submitted_at=? WHERE id=?", (now_vn().isoformat(), v["id"]))
    db.commit()
    db.close()
    flash("Đã thêm dòng báo giá thiết bị.", "success")
    return redirect(url_for("form"))


@app.route("/form/item/delete/<int:item_id>", methods=["POST"])
def item_delete(item_id):
    v = require_vendor()
    if not v:
        return redirect(url_for("dangky"))
    if window_status() == "after":
        flash("Đã hết hạn, không thể chỉnh sửa.", "danger")
        return redirect(url_for("form"))
    db = get_db()
    db.execute("DELETE FROM quote_items WHERE id=? AND vendor_id=?", (item_id, v["id"]))
    db.commit()
    db.close()
    return redirect(url_for("form"))


@app.route("/form/maint/add", methods=["POST"])
def maint_add():
    v = require_vendor()
    if not v:
        return redirect(url_for("dangky"))
    if window_status() == "after":
        flash("Đã hết hạn, không thể chỉnh sửa.", "danger")
        return redirect(url_for("form"))

    ma = request.form.get("ma_danh_muc")
    if ma not in CATEGORY_BY_MA:
        flash("Danh mục không hợp lệ.", "danger")
        return redirect(url_for("form"))

    raw = request.form.get("don_gia_baotri", "").replace(",", "").replace(".", "").strip()
    try:
        gia = int(raw) if raw else None
    except ValueError:
        gia = None

    db = get_db()
    db.execute(
        """INSERT INTO maintenance_items (vendor_id, ma_danh_muc, don_gia_baotri, ghi_chu, created_at)
           VALUES (?,?,?,?,?)""",
        (v["id"], ma, gia, request.form.get("ghi_chu", "").strip(), now_vn().isoformat()),
    )
    db.commit()
    db.close()
    flash("Đã thêm giá bảo trì/bảo hành mở rộng.", "success")
    return redirect(url_for("form"))


@app.route("/form/maint/delete/<int:item_id>", methods=["POST"])
def maint_delete(item_id):
    v = require_vendor()
    if not v:
        return redirect(url_for("dangky"))
    if window_status() == "after":
        flash("Đã hết hạn, không thể chỉnh sửa.", "danger")
        return redirect(url_for("form"))
    db = get_db()
    db.execute("DELETE FROM maintenance_items WHERE id=? AND vendor_id=?", (item_id, v["id"]))
    db.commit()
    db.close()
    return redirect(url_for("form"))


@app.route("/form/upload", methods=["POST"])
def upload():
    v = require_vendor()
    if not v:
        return redirect(url_for("dangky"))
    if window_status() == "after":
        flash("Đã hết hạn, không thể tải file.", "danger")
        return redirect(url_for("form"))

    f = request.files.get("file")
    tag = request.form.get("tag", "khac")
    if not f or f.filename == "":
        flash("Vui lòng chọn file.", "danger")
        return redirect(url_for("form"))
    if not allowed_file(f.filename):
        flash("Định dạng file không được hỗ trợ.", "danger")
        return redirect(url_for("form"))

    original = secure_filename(f.filename)
    stored = f"{v['id']}_{uuid.uuid4().hex}_{original}"
    vendor_dir = os.path.join(UPLOAD_DIR, str(v["id"]))
    os.makedirs(vendor_dir, exist_ok=True)
    f.save(os.path.join(vendor_dir, stored))

    db = get_db()
    db.execute(
        "INSERT INTO files (vendor_id, tag, original_filename, stored_filename, uploaded_at) VALUES (?,?,?,?,?)",
        (v["id"], tag, original, stored, now_vn().isoformat()),
    )
    db.commit()
    db.close()
    flash("Đã tải file lên.", "success")
    return redirect(url_for("form"))


@app.route("/form/file/delete/<int:file_id>", methods=["POST"])
def file_delete(file_id):
    v = require_vendor()
    if not v:
        return redirect(url_for("dangky"))
    if window_status() == "after":
        flash("Đã hết hạn, không thể xoá file.", "danger")
        return redirect(url_for("form"))
    db = get_db()
    row = db.execute("SELECT * FROM files WHERE id=? AND vendor_id=?", (file_id, v["id"])).fetchone()
    if row:
        try:
            os.remove(os.path.join(UPLOAD_DIR, str(v["id"]), row["stored_filename"]))
        except OSError:
            pass
        db.execute("DELETE FROM files WHERE id=?", (file_id,))
        db.commit()
    db.close()
    return redirect(url_for("form"))


@app.route("/form/file/<int:file_id>")
def file_download_own(file_id):
    v = require_vendor()
    if not v:
        return redirect(url_for("dangky"))
    db = get_db()
    row = db.execute("SELECT * FROM files WHERE id=? AND vendor_id=?", (file_id, v["id"])).fetchone()
    db.close()
    if not row:
        abort(404)
    return send_from_directory(os.path.join(UPLOAD_DIR, str(v["id"])), row["stored_filename"], as_attachment=True, download_name=row["original_filename"])
from flask import send_file
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment

@app.route("/form/export-quote")
def export_vendor_quote():
    v = require_vendor()
    if not v:
        return redirect(url_for("dangky"))

    # Đọc chính xác file mẫu in báo giá hoàn chỉnh
    template_path = os.path.join(BASE_DIR, "static", "Mau_inbaogia.xlsx")
    if not os.path.exists(template_path):
        flash("Không tìm thấy file mẫu Mau_inbaogia.xlsx trong thư mục static!", "danger")
        return redirect(url_for("form"))

    db = get_db()
    items = db.execute(
        "SELECT * FROM quote_items WHERE vendor_id=? ORDER BY id", (v["id"],)
    ).fetchall()
    db.close()

    wb = openpyxl.load_workbook(template_path)
    ws = wb["Phụ lục I Báo giá"] if "Phụ lục I Báo giá" in wb.sheetnames else wb.worksheets[0]

    # 1. Điền thông tin nhà cung cấp vào ô B5 -> B10
    ws["B5"] = v["ten_ncc"] or ""
    ws["B6"] = v["dia_chi"] or ""
    ws["B7"] = v["mst"] or ""
    ws["B8"] = v["nguoi_lien_he"] or ""
    ws["B9"] = v["sdt"] or ""
    ws["B10"] = v["email"] or ""

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    font_cell = Font(name="Times New Roman", size=11)

    # 2. Điền dữ liệu các thiết bị vào bảng từ DÒNG 14
    start_row = 14
    for idx, item in enumerate(items):
        r = start_row + idx
        ma_code = item["ma_danh_muc"]
        cat_info = CATEGORY_BY_MA.get(ma_code, {})

        # Định dạng mã có dấu cách (ví dụ TB01 -> TB 01 để khớp với sheet Bảng dữ liệu)
        formatted_ma = ma_code
        if len(ma_code) == 4 and ma_code.startswith("TB"):
            formatted_ma = f"TB {ma_code[2:]}"

        ws.cell(row=r, column=1, value=formatted_ma)                                      # Cột A: STT trong YCBG
        ws.cell(row=r, column=2, value=cat_info.get("ten", item["ten_thuong_mai"] or "")) # Cột B: Danh mục TBYT
        ws.cell(row=r, column=3, value=item["model"] or "")                               # Cột C: Model
        ws.cell(row=r, column=4, value=item["hang_sx"] or "")                             # Cột D: Hãng SX
        ws.cell(row=r, column=5, value="2026")                                            # Cột E: Năm SX
        ws.cell(row=r, column=6, value=item["xuat_xu"] or "")                             # Cột F: Xuất xứ
        ws.cell(row=r, column=7, value=cat_info.get("sl", ""))                            # Cột G: Số lượng
        ws.cell(row=r, column=8, value=cat_info.get("dvt", ""))                           # Cột H: Đơn vị tính
        ws.cell(row=r, column=9, value=item["don_gia"])                                   # Cột I: Đơn giá

        # Cột J: Thành tiền = Số lượng * Đơn giá
        sl_val = cat_info.get("sl", 0)
        dg_val = item["don_gia"] or 0
        if sl_val and dg_val:
            ws.cell(row=r, column=10, value=sl_val * dg_val)
        else:
            ws.cell(row=r, column=10, value="")

        # Gán viền, font chữ và định dạng số phân cách hàng nghìn
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            cell.font = font_cell
            cell.border = thin_border
            if c in [7, 9, 10] and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"

    # 3. Xuất file về máy
    out_io = io.BytesIO()
    wb.save(out_io)
    out_io.seek(0)

    filename = f"Bao_gia_{v['mst']}_{now_vn().strftime('%Y%m%d')}.xlsx"
    return send_file(
        out_io,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
@app.route("/thanh-cong")
def thanh_cong():
    v = require_vendor()
    if not v:
        return redirect(url_for("dangky"))

    db = get_db()
    items = db.execute("SELECT COUNT(*) as cnt FROM quote_items WHERE vendor_id=?", (v["id"],)).fetchone()
    files = db.execute("SELECT COUNT(*) as cnt FROM files WHERE vendor_id=?", (v["id"],)).fetchone()
    vendor_row = db.execute("SELECT * FROM vendors WHERE id=?", (v["id"],)).fetchone()
    db.close()

    return render_template(
        "thanhcong.html",
        vendor=vendor_row,
        item_count=items["cnt"] if items else 0,
        file_count=files["cnt"] if files else 0,
    )
# -------------------------------------------------------------- admin area --

def admin_ok():
    return session.get("is_admin") is True


@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        if request.form.get("password") == ADMIN_PASSWORD:
            session["is_admin"] = True
            return redirect(url_for("admin_home"))
        flash("Sai mật khẩu quản trị.", "danger")
    return render_template("admin_login.html")


@app.route("/admin/logout")
def admin_logout():
    session.pop("is_admin", None)
    return redirect(url_for("admin_login"))


@app.route("/admin")
def admin_home():
    if not admin_ok():
        return redirect(url_for("admin_login"))
    db = get_db()
    vendors = db.execute("SELECT * FROM vendors ORDER BY created_at").fetchall()
    counts = {}
    for v in vendors:
        n_items = db.execute("SELECT COUNT(*) c FROM quote_items WHERE vendor_id=?", (v["id"],)).fetchone()["c"]
        n_files = db.execute("SELECT COUNT(*) c FROM files WHERE vendor_id=?", (v["id"],)).fetchone()["c"]
        counts[v["id"]] = (n_items, n_files)
    db.close()
    return render_template("admin_home.html", vendors=vendors, counts=counts)


@app.route("/admin/vendor/<int:vendor_id>")
def admin_vendor(vendor_id):
    if not admin_ok():
        return redirect(url_for("admin_login"))
    db = get_db()
    v = db.execute("SELECT * FROM vendors WHERE id=?", (vendor_id,)).fetchone()
    if not v:
        abort(404)
    items = db.execute("SELECT * FROM quote_items WHERE vendor_id=? ORDER BY id", (vendor_id,)).fetchall()
    maint = db.execute("SELECT * FROM maintenance_items WHERE vendor_id=? ORDER BY id", (vendor_id,)).fetchall()
    files = db.execute("SELECT * FROM files WHERE vendor_id=? ORDER BY id", (vendor_id,)).fetchall()
    db.close()
    return render_template(
        "admin_vendor.html", vendor=v, items=items, maint=maint, files=files, category_by_ma=CATEGORY_BY_MA
    )


@app.route("/admin/download/<int:file_id>")
def admin_download(file_id):
    if not admin_ok():
        return redirect(url_for("admin_login"))
    db = get_db()
    row = db.execute("SELECT * FROM files WHERE id=?", (file_id,)).fetchone()
    db.close()
    if not row:
        abort(404)
    return send_from_directory(
        os.path.join(UPLOAD_DIR, str(row["vendor_id"])), row["stored_filename"],
        as_attachment=True, download_name=row["original_filename"],
    )


@app.route("/admin/export.xlsx")
def admin_export():
    if not admin_ok():
        return redirect(url_for("admin_login"))
    from export_excel import build_workbook
    db = get_db()
    vendors = db.execute("SELECT * FROM vendors ORDER BY created_at").fetchall()
    all_items = db.execute("SELECT * FROM quote_items ORDER BY vendor_id, id").fetchall()
    all_maint = db.execute("SELECT * FROM maintenance_items ORDER BY vendor_id, id").fetchall()
    db.close()

    vendor_by_id = {v["id"]: v for v in vendors}
    buf = io.BytesIO()
    wb = build_workbook(vendor_by_id, all_items, all_maint, CATEGORIES)
    wb.save(buf)
    buf.seek(0)
    fname = f"Tong_hop_bao_gia_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return app.response_class(
        buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@app.errorhandler(413)
def too_large(e):
    flash(f"File vượt quá dung lượng cho phép ({MAX_CONTENT_LENGTH_MB}MB).", "danger")
    return redirect(url_for("form"))


if __name__ == "__main__":
    init_db()
    app.run(debug=True, port=int(os.environ.get("PORT", 5000)))
